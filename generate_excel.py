#!/usr/bin/env python3
"""Generate an Excel workbook from the Creative Confidence self-evaluation HTML."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from html.parser import HTMLParser
import re


class QuestionParser(HTMLParser):
    """Extract questions and reverse-scored flags from the HTML."""

    def __init__(self):
        super().__init__()
        self.questions = {}  # {section: [(row_num, text, is_reverse)]}
        self.current_section = None
        self.current_row = None
        self.is_reverse = False
        self.in_td = False
        self.td_text = ""
        self.in_thead = False

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)
        if tag == "h2":
            self.current_section = None
        if tag == "table":
            id_ = attrs_dict.get("id", "")
            if id_ == "table-s1":
                self.current_section = 1
                self.questions[1] = []
            elif id_ == "table-s2":
                self.current_section = 2
                self.questions[2] = []
        if tag == "thead":
            self.in_thead = True
        if tag == "tbody":
            self.in_thead = False
        if tag == "tr" and not self.in_thead and self.current_section:
            classes = attrs_dict.get("class", "")
            self.is_reverse = "reverse" in classes
            row_num = attrs_dict.get("data-row")
            self.current_row = int(row_num) if row_num else None
        if tag == "td" and self.current_row:
            self.in_td = True
            self.td_text = ""

    def handle_endtag(self, tag):
        if tag == "td" and self.in_td and self.current_row and self.current_section:
            text = self.td_text.strip()
            if text:
                self.questions[self.current_section].append(
                    (self.current_row, text, self.is_reverse)
                )
                self.current_row = None  # only capture first td
            self.in_td = False

    def handle_data(self, data):
        if self.in_td:
            self.td_text += data


LIKERT_HEADERS = [
    "1 - Strongly disagree",
    "2 - Disagree",
    "3 - Neutral",
    "4 - Agree",
    "5 - Strongly agree",
]

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
REVERSE_FILL = PatternFill("solid", fgColor="FFF5D6")
SECTION_FILL = PatternFill("solid", fgColor="27AE60")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def build_workbook(questions: dict) -> Workbook:
    wb = Workbook()

    # --- Sheet 1: Questionnaire (answer entry) ---
    ws = wb.active
    ws.title = "Questionnaire"

    # Title row
    ws.merge_cells("A1:G1")
    ws["A1"] = "Auto-evaluation de ma Confiance creative"
    ws["A1"].font = Font(bold=True, size=14, color="2C3E50")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Metadata rows
    ws["A3"] = "Nom / Name:"
    ws["A3"].font = Font(bold=True)
    ws.merge_cells("B3:G3")
    ws["B3"].border = Border(bottom=Side(style="thin"))

    ws["A4"] = "Date:"
    ws["A4"].font = Font(bold=True)
    ws.merge_cells("B4:G4")
    ws["B4"].border = Border(bottom=Side(style="thin"))

    row = 6

    for section_num in [1, 2]:
        # Section header
        ws.merge_cells(f"A{row}:G{row}")
        cell = ws.cell(row=row, column=1, value=f"Section {section_num}")
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.alignment = Alignment(horizontal="center")
        row += 1

        # Column headers
        headers = ["#", "Enonce", *LIKERT_HEADERS]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER
        row += 1

        # Questions
        for q_row, text, is_reverse in questions[section_num]:
            ws.cell(row=row, column=1, value=q_row).alignment = Alignment(
                horizontal="center"
            )
            ws.cell(row=row, column=1).border = THIN_BORDER

            label = f"{text}  (reverse)" if is_reverse else text
            cell_q = ws.cell(row=row, column=2, value=label)
            cell_q.border = THIN_BORDER
            cell_q.alignment = Alignment(wrap_text=True)

            if is_reverse:
                for c in range(1, 8):
                    ws.cell(row=row, column=c).fill = REVERSE_FILL

            # Empty cells for Likert responses (1-5)
            for col in range(3, 8):
                ws.cell(row=row, column=col).border = THIN_BORDER
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal="center"
                )

            row += 1

        # Section subtotal row
        ws.cell(row=row, column=1, value="").border = THIN_BORDER
        cell_sub = ws.cell(row=row, column=2, value=f"Score Section {section_num}")
        cell_sub.font = Font(bold=True)
        cell_sub.border = THIN_BORDER
        for col in range(3, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        # Formula: sum columns C-G for this section's rows
        start = row - len(questions[section_num])
        end = row - 1
        ws.merge_cells(f"C{row}:G{row}")
        ws.cell(row=row, column=3).value = f"=SUM(C{start}:G{end})"
        ws.cell(row=row, column=3).font = Font(bold=True, color="27AE60", size=14)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3).border = THIN_BORDER
        row += 2

    # Global score
    ws.cell(row=row, column=2, value="Score Global (max 200)").font = Font(
        bold=True, size=12
    )
    row += 1
    row += 1

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 70
    for col in range(3, 8):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # --- Sheet 2: Scoring Guide ---
    ws2 = wb.create_sheet("Scoring Guide")
    ws2["A1"] = "Scoring Guide - Auto-evaluation Confiance creative"
    ws2["A1"].font = Font(bold=True, size=14)

    ws2["A3"] = "Standard items: score = response value (1-5)"
    ws2["A4"] = "Reverse items (marked in yellow): score = 6 - response value"
    ws2["A6"] = "Reverse-scored rows per section:"
    ws2["A7"] = "  Row 3, 6, 11, 17"
    ws2["A9"] = "Score ranges:"
    ws2["A10"] = "  Per section (20 items): 20-100"
    ws2["A11"] = "  Global (40 items): 40-200"
    ws2["A13"] = "Higher scores = greater creative confidence"

    # --- Sheet 3: Data Dictionary ---
    ws3 = wb.create_sheet("Data Dictionary")
    headers = ["Item #", "Section", "Enonce", "Reverse?", "Scoring"]
    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    r = 2
    for section in [1, 2]:
        for q_row, text, is_rev in questions[section]:
            ws3.cell(row=r, column=1, value=q_row).border = THIN_BORDER
            ws3.cell(row=r, column=2, value=section).border = THIN_BORDER
            ws3.cell(row=r, column=3, value=text).border = THIN_BORDER
            ws3.cell(row=r, column=4, value="Yes" if is_rev else "No").border = THIN_BORDER
            ws3.cell(row=r, column=5, value="6 - val" if is_rev else "val").border = THIN_BORDER
            if is_rev:
                for c in range(1, 6):
                    ws3.cell(row=r, column=c).fill = REVERSE_FILL
            r += 1

    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 80
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 12

    return wb


def main():
    html_path = Path(__file__).parent / "Auto-évaluation de ma Confiance créative.html"
    output_path = Path(__file__).parent / "Auto-evaluation_Confiance_creative.xlsx"

    html_content = html_path.read_text(encoding="utf-8")
    parser = QuestionParser()
    parser.feed(html_content)

    if not parser.questions.get(1) or not parser.questions.get(2):
        print("ERROR: Could not parse questions from HTML.")
        return

    print(f"Parsed {len(parser.questions[1])} questions (Section 1), "
          f"{len(parser.questions[2])} questions (Section 2)")

    wb = build_workbook(parser.questions)
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
